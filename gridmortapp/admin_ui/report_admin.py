# gridmortapp/admin_ui/report_admin.py
from django.contrib import admin
from django.contrib.auth.models import User
from django.db.models import Q, Count, Avg, F
from django.db import IntegrityError
from django.utils import timezone
from django.utils.html import mark_safe
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
import csv
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from unfold.admin import ModelAdmin
from unfold.decorators import display

from gridmortapp.models import ReportType, Report, ReportLog
from gridmortapp.system_models.ticket_models import Ticket, TicketStatus, TicketPriority
from gridmortapp.models import HardwareItem, Department, EmployeeProfile


@admin.register(ReportType)
class ReportTypeAdmin(ModelAdmin):
    list_display = ['name', 'report_code', 'report_count', 'is_active', 'created_at']
    search_fields = ['name', 'report_code']
    list_filter = ['is_active']
    
    def report_count(self, obj):
        return obj.report_set.count()
    report_count.short_description = "Reports"


@admin.register(ReportLog)
class ReportLogAdmin(ModelAdmin):
    list_display = ['report', 'user', 'action', 'timestamp']
    list_filter = ['action', 'timestamp']
    search_fields = ['report__title', 'user__username']
    readonly_fields = ['timestamp']
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False


@admin.register(Report)
class ReportAdmin(ModelAdmin):
    list_display = ['title', 'report_type', 'status_display', 'generated_at', 'generated_by']
    list_filter = ['report_type', 'status', 'generated_at']
    search_fields = ['title', 'description']
    readonly_fields = ['generated_at', 'file_size']
    date_hierarchy = 'generated_at'
    
    fieldsets = (
        ('Report Information', {
            'fields': ('title', 'description', 'report_type', 'status')
        }),
        ('Statistics', {
            'fields': ('total_tickets', 'resolved_tickets', 'unresolved_tickets', 'avg_resolution_time', 
                      'total_hardware', 'assigned_hardware', 'available_hardware', 'maintenance_hardware')
        }),
        ('Generation', {
            'fields': ('generated_by', 'generated_at', 'file_size')
        }),
        ('Additional', {
            'fields': ('notes',)
        }),
    )
    
    def changelist_view(self, request, extra_context=None):
        """Auto-generate report entries when visiting the page"""
        self.ensure_report_types_exist()
        self.auto_generate_report_entries(request)
        return super().changelist_view(request, extra_context)
    
    def ensure_report_types_exist(self):
        """Ensure all required ReportTypes exist"""
        report_types = [
            {'code': 'TICKET_SUMMARY', 'name': 'Ticket Summary Report'},
            {'code': 'HARDWARE_SUMMARY', 'name': 'Hardware Inventory Report'},
            {'code': 'AGENT_PERFORMANCE', 'name': 'Agent Performance Report'},
            {'code': 'DEPARTMENT_SUMMARY', 'name': 'Department Summary Report'},
            {'code': 'SLA_COMPLIANCE', 'name': 'SLA Compliance Report'},
        ]
        
        for rt in report_types:
            try:
                ReportType.objects.get_or_create(
                    report_code=rt['code'],
                    defaults={'name': rt['name'], 'is_active': True}
                )
            except IntegrityError:
                pass
    
    def auto_generate_report_entries(self, request):
        """Auto-create report entries for today if they don't exist"""
        today = timezone.now().date()
        
        existing_reports = Report.objects.filter(
            generated_at__date=today
        ).values_list('report_type__report_code', flat=True)
        existing_codes = list(existing_reports)
        
        user = request.user if request.user.is_authenticated else None
        
        # Only create the report entries, don't populate with stale data
        if 'TICKET_SUMMARY' not in existing_codes:
            self.create_report_entry('TICKET_SUMMARY', 'Ticket Summary Report', user)
        
        if 'HARDWARE_SUMMARY' not in existing_codes:
            self.create_report_entry('HARDWARE_SUMMARY', 'Hardware Inventory Report', user)
        
        if 'AGENT_PERFORMANCE' not in existing_codes:
            self.create_report_entry('AGENT_PERFORMANCE', 'Agent Performance Report', user)
        
        if 'DEPARTMENT_SUMMARY' not in existing_codes:
            self.create_report_entry('DEPARTMENT_SUMMARY', 'Department Summary Report', user)
        
        if 'SLA_COMPLIANCE' not in existing_codes:
            self.create_report_entry('SLA_COMPLIANCE', 'SLA Compliance Report', user)
    
    def create_report_entry(self, report_code, report_name, user):
        """Create a report entry without populating statistics (they'll be fetched on download)"""
        report_type = ReportType.objects.get(report_code=report_code)
        
        report = Report.objects.create(
            title=f"{report_name} - {timezone.now().strftime('%Y-%m-%d')}",
            description=f"Auto-generated {report_name.lower()}",
            report_type=report_type,
            status='generated',
            generated_by=user,
            # Leave statistics as 0 - they'll be populated on download
            total_tickets=0,
            resolved_tickets=0,
            unresolved_tickets=0,
            avg_resolution_time=0,
            total_hardware=0,
            assigned_hardware=0,
            available_hardware=0,
            maintenance_hardware=0,
            notes="Report will fetch fresh data on download"
        )
        
        ReportLog.objects.create(
            report=report,
            user=user,
            action='generated',
            details={'auto_generated': True}
        )
    
    def get_fresh_report_data(self, report_code):
        """Fetch fresh data from database for a specific report type"""
        if report_code == 'TICKET_SUMMARY':
            return self.get_ticket_summary_data()
        elif report_code == 'HARDWARE_SUMMARY':
            return self.get_hardware_summary_data()
        elif report_code == 'AGENT_PERFORMANCE':
            return self.get_agent_performance_data()
        elif report_code == 'DEPARTMENT_SUMMARY':
            return self.get_department_summary_data()
        elif report_code == 'SLA_COMPLIANCE':
            return self.get_sla_compliance_data()
        return {}
    
    def get_ticket_summary_data(self):
        """Fetch fresh ticket data"""
        tickets = Ticket.objects.all()
        total = tickets.count()
        resolved = tickets.filter(status__status_type='resolved').count()
        open_tickets = tickets.filter(status__status_type__in=['new', 'open', 'in_progress']).count()
        closed = tickets.filter(status__status_type='closed').count()
        sla_breached = tickets.filter(sla_breached=True).count()
        
        avg_res = tickets.filter(resolved_at__isnull=False).aggregate(
            avg=Avg(F('resolved_at') - F('created_at'))
        )
        avg_hours = round(avg_res['avg'].total_seconds() / 3600, 1) if avg_res['avg'] else 0
        
        status_breakdown = {}
        for item in tickets.values('status__name').annotate(count=Count('id')):
            status_breakdown[item['status__name']] = item['count']
        
        priority_breakdown = {}
        for item in tickets.values('priority__name').annotate(count=Count('id')):
            priority_breakdown[item['priority__name']] = item['count']
        
        return {
            'total': total,
            'resolved': resolved,
            'unresolved': open_tickets,
            'closed': closed,
            'sla_breached': sla_breached,
            'avg_hours': avg_hours,
            'status_breakdown': status_breakdown,
            'priority_breakdown': priority_breakdown
        }
    
    def get_hardware_summary_data(self):
        """Fetch fresh hardware data"""
        hardware = HardwareItem.objects.all()
        total = hardware.count()
        available = hardware.filter(status='available').count()
        assigned = hardware.filter(status='assigned').count()
        maintenance = hardware.filter(status='maintenance').count()
        repair = hardware.filter(status='repair').count()
        retired = hardware.filter(status='retired').count()
        lost = hardware.filter(status='lost').count()
        
        category_breakdown = {}
        for item in hardware.values('category__name').annotate(count=Count('id')):
            category_breakdown[item['category__name']] = item['count']
        
        condition_breakdown = {}
        for item in hardware.values('condition').annotate(count=Count('id')):
            condition_breakdown[item['condition']] = item['count']
        
        return {
            'total': total,
            'available': available,
            'assigned': assigned,
            'maintenance': maintenance,
            'repair': repair,
            'retired': retired,
            'lost': lost,
            'category_breakdown': category_breakdown,
            'condition_breakdown': condition_breakdown
        }
    
    def get_agent_performance_data(self):
        """Fetch fresh agent performance data"""
        agents = User.objects.filter(Q(groups__name='IT Staff') | Q(groups__name='Manager')).distinct()
        agent_data = []
        
        for agent in agents:
            assigned_tickets = Ticket.objects.filter(assignee=agent)
            total_assigned = assigned_tickets.count()
            resolved = assigned_tickets.filter(status__status_type='resolved').count()
            open_tickets = assigned_tickets.filter(status__status_type__in=['new', 'open', 'in_progress']).count()
            rate = round((resolved / total_assigned * 100), 1) if total_assigned > 0 else 0
            
            agent_data.append({
                'name': agent.get_full_name() or agent.username,
                'assigned': total_assigned,
                'resolved': resolved,
                'open': open_tickets,
                'completion_rate': rate
            })
        
        return agent_data
    
    def get_department_summary_data(self):
        """Fetch fresh department data"""
        departments = Department.objects.all()
        dept_data = []
        
        for dept in departments:
            users = EmployeeProfile.objects.filter(department=dept).values_list('user', flat=True)
            dept_tickets = Ticket.objects.filter(requestor__in=users)
            total = dept_tickets.count()
            open_dept = dept_tickets.filter(status__status_type__in=['new', 'open', 'in_progress']).count()
            resolved = dept_tickets.filter(status__status_type='resolved').count()
            
            dept_data.append({
                'name': dept.name,
                'total': total,
                'open': open_dept,
                'resolved': resolved
            })
        
        return dept_data
    
    def get_sla_compliance_data(self):
        """Fetch fresh SLA data"""
        tickets = Ticket.objects.all()
        total_sla = tickets.filter(priority__isnull=False).count()
        breached = tickets.filter(sla_breached=True).count()
        compliant = total_sla - breached
        compliance_rate = round((compliant / total_sla * 100), 1) if total_sla > 0 else 0
        
        priority_data = []
        for priority in TicketPriority.objects.filter(is_active=True):
            tickets_with_priority = tickets.filter(priority=priority)
            total = tickets_with_priority.count()
            breached_count = tickets_with_priority.filter(sla_breached=True).count()
            met = total - breached_count
            rate = round((met / total * 100), 1) if total > 0 else 0
            
            priority_data.append({
                'name': priority.name,
                'total': total,
                'compliant': met,
                'breached': breached_count,
                'compliance_rate': rate
            })
        
        return {
            'total_sla': total_sla,
            'compliant': compliant,
            'breached': breached,
            'compliance_rate': compliance_rate,
            'priority_data': priority_data
        }
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        is_authorized = (
            request.user.is_superuser or 
            request.user.groups.filter(name__in=['IT Staff', 'Support Agents', 'Manager', 'Admin']).exists()
        )
        if not is_authorized:
            return qs.filter(generated_by=request.user)
        return qs
    
    @display(description="Status")
    def status_display(self, obj):
        colors = {
            'draft': '#6B7280',
            'generated': '#10B981',
            'scheduled': '#3B82F6',
            'archived': '#F59E0B',
        }
        color = colors.get(obj.status, '#808080')
        return mark_safe(f'<span style="color: {color}; font-weight: bold;">{obj.get_status_display()}</span>')
    
    actions = ['download_reports_action', 'archive_reports', 'delete_reports']
    
    def download_reports_action(self, request, queryset):
        """Download selected reports with fresh data"""
        if not queryset:
            self.message_user(request, "No reports selected", level='ERROR')
            return
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Style definitions
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create a sheet for each report with fresh data
        for i, report in enumerate(queryset):
            if i == 0:
                ws = wb.active
                ws.title = report.report_type.report_code[:30] if report.report_type else f"Report_{i+1}"
            else:
                sheet_name = report.report_type.report_code[:30] if report.report_type else f"Report_{i+1}"
                ws = wb.create_sheet(sheet_name)
            
            # Get fresh data based on report type
            report_code = report.report_type.report_code if report.report_type else None
            
            if report_code == 'TICKET_SUMMARY':
                self.populate_ticket_summary_sheet(ws, report)
            elif report_code == 'HARDWARE_SUMMARY':
                self.populate_hardware_summary_sheet(ws, report)
            elif report_code == 'AGENT_PERFORMANCE':
                self.populate_agent_performance_sheet(ws, report)
            elif report_code == 'DEPARTMENT_SUMMARY':
                self.populate_department_summary_sheet(ws, report)
            elif report_code == 'SLA_COMPLIANCE':
                self.populate_sla_compliance_sheet(ws, report)
            else:
                ws['A1'] = f"Report: {report.title}"
                ws['A2'] = "No data available for this report type"
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Log downloads
        for report in queryset:
            ReportLog.objects.create(
                report=report,
                user=request.user if request.user.is_authenticated else None,
                action='downloaded',
                details={'format': 'excel', 'bulk_download': True, 'fresh_data': True}
            )
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reports_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb.save(response)
        return response
    
    download_reports_action.short_description = "Download selected reports as Excel"
    
    def populate_ticket_summary_sheet(self, ws, report):
        """Populate ticket summary sheet with fresh data"""
        data = self.get_ticket_summary_data()
        
        row = 1
        ws.merge_cells(f'A{i+1}:D{i+1}')
        ws[f'A{row}'] = report.title
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2
        
        # Metadata
        ws[f'A{row}'] = 'Generated By:'
        ws[f'B{row}'] = report.generated_by.get_full_name() if report.generated_by else 'System'
        row += 1
        ws[f'A{row}'] = 'Generated At:'
        ws[f'B{row}'] = timezone.now().strftime('%Y-%m-%d %H:%M')
        row += 1
        ws[f'A{row}'] = 'Data Freshness:'
        ws[f'B{row}'] = 'Live data from database'
        row += 2
        
        # Main statistics
        ws[f'A{row}'] = 'TICKET STATISTICS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'Total Tickets'
        ws[f'B{row}'] = data['total']
        row += 1
        ws[f'A{row}'] = 'Resolved Tickets'
        ws[f'B{row}'] = data['resolved']
        row += 1
        ws[f'A{row}'] = 'Unresolved Tickets'
        ws[f'B{row}'] = data['unresolved']
        row += 1
        ws[f'A{row}'] = 'Closed Tickets'
        ws[f'B{row}'] = data['closed']
        row += 1
        ws[f'A{row}'] = 'SLA Breached'
        ws[f'B{row}'] = data['sla_breached']
        row += 1
        ws[f'A{row}'] = 'Avg Resolution Time (hours)'
        ws[f'B{row}'] = data['avg_hours']
        row += 2
        
        # Status breakdown
        if data['status_breakdown']:
            ws[f'A{row}'] = 'BY STATUS'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            for status, count in data['status_breakdown'].items():
                ws[f'A{row}'] = status
                ws[f'B{row}'] = count
                row += 1
            row += 1
        
        # Priority breakdown
        if data['priority_breakdown']:
            ws[f'A{row}'] = 'BY PRIORITY'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            for priority, count in data['priority_breakdown'].items():
                ws[f'A{row}'] = priority
                ws[f'B{row}'] = count
                row += 1
    
    def populate_hardware_summary_sheet(self, ws, report):
        """Populate hardware summary sheet with fresh data"""
        data = self.get_hardware_summary_data()
        
        row = 1
        ws.merge_cells(f'A{i+1}:D{i+1}')
        ws[f'A{row}'] = report.title
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2
        
        ws[f'A{row}'] = 'Generated By:'
        ws[f'B{row}'] = report.generated_by.get_full_name() if report.generated_by else 'System'
        row += 1
        ws[f'A{row}'] = 'Generated At:'
        ws[f'B{row}'] = timezone.now().strftime('%Y-%m-%d %H:%M')
        row += 1
        ws[f'A{row}'] = 'Data Freshness:'
        ws[f'B{row}'] = 'Live data from database'
        row += 2
        
        ws[f'A{row}'] = 'HARDWARE STATISTICS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'Total Hardware'
        ws[f'B{row}'] = data['total']
        row += 1
        ws[f'A{row}'] = 'Available'
        ws[f'B{row}'] = data['available']
        row += 1
        ws[f'A{row}'] = 'Assigned'
        ws[f'B{row}'] = data['assigned']
        row += 1
        ws[f'A{row}'] = 'Maintenance'
        ws[f'B{row}'] = data['maintenance']
        row += 1
        ws[f'A{row}'] = 'Repair'
        ws[f'B{row}'] = data['repair']
        row += 1
        ws[f'A{row}'] = 'Retired'
        ws[f'B{row}'] = data['retired']
        row += 1
        ws[f'A{row}'] = 'Lost'
        ws[f'B{row}'] = data['lost']
        row += 2
        
        if data['category_breakdown']:
            ws[f'A{row}'] = 'BY CATEGORY'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            for category, count in data['category_breakdown'].items():
                ws[f'A{row}'] = category
                ws[f'B{row}'] = count
                row += 1
            row += 1
        
        if data['condition_breakdown']:
            ws[f'A{row}'] = 'BY CONDITION'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            for condition, count in data['condition_breakdown'].items():
                ws[f'A{row}'] = condition
                ws[f'B{row}'] = count
                row += 1
    
    def populate_agent_performance_sheet(self, ws, report):
        """Populate agent performance sheet with fresh data"""
        data = self.get_agent_performance_data()
        
        row = 1
        ws.merge_cells(f'A{i+1}:D{i+1}')
        ws[f'A{row}'] = report.title
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2
        
        ws[f'A{row}'] = 'Generated By:'
        ws[f'B{row}'] = report.generated_by.get_full_name() if report.generated_by else 'System'
        row += 1
        ws[f'A{row}'] = 'Generated At:'
        ws[f'B{row}'] = timezone.now().strftime('%Y-%m-%d %H:%M')
        row += 1
        ws[f'A{row}'] = 'Data Freshness:'
        ws[f'B{row}'] = 'Live data from database'
        row += 2
        
        ws[f'A{row}'] = 'AGENT PERFORMANCE'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers
        ws[f'A{row}'] = 'Agent'
        ws[f'B{row}'] = 'Assigned'
        ws[f'C{row}'] = 'Resolved'
        ws[f'D{row}'] = 'Open'
        ws[f'E{row}'] = 'Completion Rate %'
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True)
        row += 1
        
        for agent in data:
            ws[f'A{row}'] = agent['name']
            ws[f'B{row}'] = agent['assigned']
            ws[f'C{row}'] = agent['resolved']
            ws[f'D{row}'] = agent['open']
            ws[f'E{row}'] = agent['completion_rate']
            row += 1
    
    def populate_department_summary_sheet(self, ws, report):
        """Populate department summary sheet with fresh data"""
        data = self.get_department_summary_data()
        
        row = 1
        ws.merge_cells(f'A{i+1}:D{i+1}')
        ws[f'A{row}'] = report.title
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2
        
        ws[f'A{row}'] = 'Generated By:'
        ws[f'B{row}'] = report.generated_by.get_full_name() if report.generated_by else 'System'
        row += 1
        ws[f'A{row}'] = 'Generated At:'
        ws[f'B{row}'] = timezone.now().strftime('%Y-%m-%d %H:%M')
        row += 1
        ws[f'A{row}'] = 'Data Freshness:'
        ws[f'B{row}'] = 'Live data from database'
        row += 2
        
        ws[f'A{row}'] = 'DEPARTMENT SUMMARY'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = 'Department'
        ws[f'B{row}'] = 'Total Tickets'
        ws[f'C{row}'] = 'Open'
        ws[f'D{row}'] = 'Resolved'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
        row += 1
        
        for dept in data:
            ws[f'A{row}'] = dept['name']
            ws[f'B{row}'] = dept['total']
            ws[f'C{row}'] = dept['open']
            ws[f'D{row}'] = dept['resolved']
            row += 1
    
    def populate_sla_compliance_sheet(self, ws, report):
        """Populate SLA compliance sheet with fresh data"""
        data = self.get_sla_compliance_data()
        
        row = 1
        ws.merge_cells(f'A{i+1}:D{i+1}')
        ws[f'A{row}'] = report.title
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2
        
        ws[f'A{row}'] = 'Generated By:'
        ws[f'B{row}'] = report.generated_by.get_full_name() if report.generated_by else 'System'
        row += 1
        ws[f'A{row}'] = 'Generated At:'
        ws[f'B{row}'] = timezone.now().strftime('%Y-%m-%d %H:%M')
        row += 1
        ws[f'A{row}'] = 'Data Freshness:'
        ws[f'B{row}'] = 'Live data from database'
        row += 2
        
        ws[f'A{row}'] = 'SLA COMPLIANCE'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = 'Total Tickets with SLA'
        ws[f'B{row}'] = data['total_sla']
        row += 1
        ws[f'A{row}'] = 'Compliant'
        ws[f'B{row}'] = data['compliant']
        row += 1
        ws[f'A{row}'] = 'Breached'
        ws[f'B{row}'] = data['breached']
        row += 1
        ws[f'A{row}'] = 'Compliance Rate'
        ws[f'B{row}'] = f"{data['compliance_rate']}%"
        row += 2
        
        if data['priority_data']:
            ws[f'A{row}'] = 'BY PRIORITY'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            ws[f'A{row}'] = 'Priority'
            ws[f'B{row}'] = 'Total'
            ws[f'C{row}'] = 'Compliant'
            ws[f'D{row}'] = 'Breached'
            ws[f'E{row}'] = 'Compliance Rate %'
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].font = Font(bold=True)
            row += 1
            
            for priority in data['priority_data']:
                ws[f'A{row}'] = priority['name']
                ws[f'B{row}'] = priority['total']
                ws[f'C{row}'] = priority['compliant']
                ws[f'D{row}'] = priority['breached']
                ws[f'E{row}'] = priority['compliance_rate']
                row += 1
    
    def archive_reports(self, request, queryset):
        count = queryset.update(status='archived')
        for report in queryset:
            ReportLog.objects.create(
                report=report,
                user=request.user,
                action='archived',
                details={'archived_by': request.user.username}
            )
        self.message_user(request, f"{count} reports archived.")
    archive_reports.short_description = "Archive selected reports"
    
    def delete_reports(self, request, queryset):
        count = queryset.count()
        for report in queryset:
            ReportLog.objects.create(
                report=report,
                user=request.user,
                action='deleted',
                details={'deleted_by': request.user.username}
            )
        queryset.delete()
        self.message_user(request, f"{count} reports deleted.")
    delete_reports.short_description = "Delete selected reports"