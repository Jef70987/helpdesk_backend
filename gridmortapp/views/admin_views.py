from rest_framework import status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth.models import User
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from datetime import timedelta
import logging

from gridmortapp.system_models.ticket_models import (
    Ticket, TicketMessage, TicketCategory, TicketStatus, TicketPriority
)
from gridmortapp.system_models.audit_models import AuditLog
from gridmortapp.models import HardwareItem, HardwareCategory, InventoryMovement, Department, EmployeeProfile
from gridmortapp.serializers.app_serializers import (
    UserSerializer, TicketListSerializer, TicketDetailSerializer,
    TicketMessageSerializer, TicketCategorySerializer
)

logger = logging.getLogger(__name__)


def is_admin(user):
    """Check if user has admin privileges"""
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=['Admin', 'Manager']).exists()


@method_decorator(csrf_exempt, name='dispatch')
class AdminDashboardView(APIView):
    """Admin dashboard statistics and analytics"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # User Statistics
        total_users = User.objects.count()
        active_users = User.objects.filter(is_active=True).count()
        
        # Ticket Statistics
        tickets = Ticket.objects.all()
        total_tickets = tickets.count()
        open_tickets = tickets.filter(
            status__status_type__in=['new', 'open', 'in_progress', 'pending_customer', 'pending_third_party']
        ).count()
        resolved_tickets = tickets.filter(status__status_type='resolved').count()
        closed_tickets = tickets.filter(status__status_type='closed').count()
        sla_breached = tickets.filter(sla_breached=True).count()
        
        # Hardware Statistics
        hardware = HardwareItem.objects.all()
        total_hardware = hardware.count()
        available_hardware = hardware.filter(status='available').count()
        assigned_hardware = hardware.filter(status='assigned').count()
        maintenance_hardware = hardware.filter(status='maintenance').count()
        
        # Ticket Trend (last 7 days)
        now = timezone.now()
        trend_data = []
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
            count = tickets.filter(created_at__gte=day_start, created_at__lte=day_end).count()
            trend_data.append({
                'day': day.strftime('%a'),
                'date': day.strftime('%Y-%m-%d'),
                'count': count
            })
        
        # Category Breakdown
        category_breakdown = []
        categories = TicketCategory.objects.filter(is_active=True)
        for cat in categories:
            count = tickets.filter(category=cat).count()
            if count > 0:
                category_breakdown.append({
                    'name': cat.name,
                    'count': count
                })
        
        # Status Breakdown
        status_breakdown = []
        statuses = TicketStatus.objects.filter(is_active=True)
        for status in statuses:
            count = tickets.filter(status=status).count()
            if count > 0:
                status_breakdown.append({
                    'name': status.status_type,
                    'count': count
                })
        
        # Hardware Status Breakdown
        hardware_status_breakdown = []
        status_list = ['available', 'assigned', 'maintenance', 'repair', 'retired', 'lost']
        for status in status_list:
            count = hardware.filter(status=status).count()
            if count > 0:
                hardware_status_breakdown.append({
                    'name': status,
                    'count': count
                })
        
        # Recent Tickets
        recent_tickets = tickets.order_by('-created_at')[:10]
        recent_tickets_data = TicketListSerializer(recent_tickets, many=True, context={'request': request}).data
        
        # Agent Performance
        agent_performance = []
        agents = User.objects.filter(groups__name__in=['IT Staff', 'Support Agents']).distinct()
        for agent in agents:
            assigned = tickets.filter(assignee=agent).count()
            resolved = tickets.filter(assignee=agent, status__status_type='resolved').count()
            open_count = tickets.filter(assignee=agent, status__status_type__in=['new', 'open', 'in_progress']).count()
            rate = round((resolved / assigned * 100), 1) if assigned > 0 else 0
            if assigned > 0:
                agent_performance.append({
                    'name': agent.get_full_name() or agent.username,
                    'assigned': assigned,
                    'resolved': resolved,
                    'open': open_count,
                    'completion_rate': rate
                })
        
        # Department Breakdown
        department_breakdown = []
        departments = Department.objects.all()
        for dept in departments:
            users = EmployeeProfile.objects.filter(department=dept).values_list('user', flat=True)
            count = tickets.filter(requestor__in=users).count()
            if count > 0:
                department_breakdown.append({
                    'name': dept.name,
                    'count': count
                })
        
        return Response({
            'stats': {
                'total_users': total_users,
                'active_users': active_users,
                'total_tickets': total_tickets,
                'open_tickets': open_tickets,
                'resolved_tickets': resolved_tickets,
                'closed_tickets': closed_tickets,
                'sla_breached': sla_breached,
                'total_hardware': total_hardware,
                'available_hardware': available_hardware,
                'assigned_hardware': assigned_hardware,
                'maintenance_hardware': maintenance_hardware,
            },
            'ticket_trend': trend_data,
            'category_breakdown': category_breakdown,
            'status_breakdown': status_breakdown,
            'hardware_status_breakdown': hardware_status_breakdown,
            'recent_tickets': recent_tickets_data,
            'agent_performance': agent_performance,
            'department_breakdown': department_breakdown,
        })


@method_decorator(csrf_exempt, name='dispatch')
class AdminTicketListView(APIView):
    """List all tickets with filters"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        status_filter = request.query_params.get('status')
        search = request.query_params.get('search')
        
        tickets = Ticket.objects.all()
        
        if status_filter:
            tickets = tickets.filter(status__status_type=status_filter)
        
        if search:
            tickets = tickets.filter(
                Q(ticket_id__icontains=search) |
                Q(title__icontains=search) |
                Q(requestor__username__icontains=search) |
                Q(requestor__email__icontains=search)
            )
        
        tickets = tickets.order_by('-created_at')
        
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        start = (page - 1) * page_size
        end = start + page_size
        
        paginated_tickets = tickets[start:end]
        serializer = TicketListSerializer(paginated_tickets, many=True, context={'request': request})
        
        return Response({
            'results': serializer.data,
            'count': tickets.count(),
            'page': page,
            'page_size': page_size
        })


@method_decorator(csrf_exempt, name='dispatch')
class AdminTicketDetailView(APIView):
    """Get, update, delete ticket"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            ticket = Ticket.objects.get(pk=pk)
            serializer = TicketDetailSerializer(ticket, context={'request': request})
            return Response(serializer.data)
        except Ticket.DoesNotExist:
            return Response(
                {'error': 'Ticket not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    def patch(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            ticket = Ticket.objects.get(pk=pk)
        except Ticket.DoesNotExist:
            return Response(
                {'error': 'Ticket not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        status_type = request.data.get('status')
        assignee_id = request.data.get('assignee')
        
        if status_type:
            try:
                status_obj = TicketStatus.objects.get(status_type=status_type)
                ticket.status = status_obj
                if status_type in ['resolved', 'closed'] and not ticket.resolved_at:
                    ticket.resolved_at = timezone.now()
                ticket.save()
            except TicketStatus.DoesNotExist:
                return Response(
                    {'error': 'Invalid status'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        if assignee_id:
            try:
                assignee = User.objects.get(pk=assignee_id)
                ticket.assignee = assignee
                if not ticket.first_assignment_at:
                    ticket.first_assignment_at = timezone.now()
                ticket.save()
            except User.DoesNotExist:
                return Response(
                    {'error': 'Invalid assignee'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        AuditLog.objects.create(
            ticket=ticket,
            user=request.user,
            action='TICKET_UPDATED',
            details={
                'ticket_id': ticket.ticket_id,
                'updated_by': request.user.username
            }
        )
        
        serializer = TicketDetailSerializer(ticket, context={'request': request})
        return Response(serializer.data)

    def delete(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            ticket = Ticket.objects.get(pk=pk)
        except Ticket.DoesNotExist:
            return Response(
                {'error': 'Ticket not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        AuditLog.objects.create(
            ticket=ticket,
            user=request.user,
            action='TICKET_DELETED',
            details={
                'ticket_id': ticket.ticket_id,
                'deleted_by': request.user.username
            }
        )
        
        ticket.delete()
        return Response({'success': True, 'message': 'Ticket deleted successfully'})


@method_decorator(csrf_exempt, name='dispatch')
class AdminTicketMessageView(APIView):
    """Get messages and send messages for any ticket"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            ticket = Ticket.objects.get(pk=pk)
            messages = ticket.messages.all().order_by('created_at')
            serializer = TicketMessageSerializer(messages, many=True, context={'request': request})
            return Response(serializer.data)
        except Ticket.DoesNotExist:
            return Response(
                {'error': 'Ticket not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            ticket = Ticket.objects.get(pk=pk)
        except Ticket.DoesNotExist:
            return Response(
                {'error': 'Ticket not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        content = request.data.get('content')
        if not content:
            return Response(
                {'error': 'Message content is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        message = TicketMessage.objects.create(
            ticket=ticket,
            author=request.user,
            content=content,
            message_type='public'
        )

        ticket.last_modified_at = timezone.now()
        ticket.save()

        AuditLog.objects.create(
            ticket=ticket,
            user=request.user,
            action='MESSAGE_ADDED',
            details={
                'ticket_id': ticket.ticket_id,
                'message_id': message.id,
                'sent_by': request.user.username
            }
        )

        serializer = TicketMessageSerializer(message, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name='dispatch')
class AdminUserListView(APIView):
    """List all users"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        users = User.objects.all().order_by('-date_joined')
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data)


@method_decorator(csrf_exempt, name='dispatch')
class AdminUserDetailView(APIView):
    """Get, update, delete user"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            user = User.objects.get(pk=pk)
            serializer = UserSerializer(user)
            return Response(serializer.data)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    def patch(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        data = request.data
        if 'first_name' in data:
            user.first_name = data['first_name']
        if 'last_name' in data:
            user.last_name = data['last_name']
        if 'email' in data:
            user.email = data['email']
        if 'username' in data:
            user.username = data['username']
        if 'is_active' in data:
            user.is_active = data['is_active']
        user.save()

        # Update profile if fields provided
        try:
            profile = user.profile
            if 'user_type' in data:
                profile.user_type = data['user_type']
            if 'department' in data:
                from gridmortapp.models import Department
                try:
                    dept = Department.objects.get(name=data['department'])
                    profile.department = dept
                except Department.DoesNotExist:
                    pass
            if 'employee_id' in data:
                profile.employee_id = data['employee_id']
            if 'position' in data:
                profile.position = data['position']
            if 'phone_number' in data:
                profile.phone_number = data['phone_number']
            if 'office_location' in data:
                profile.office_location = data['office_location']
            profile.save()
        except:
            pass

        AuditLog.objects.create(
            user=request.user,
            action='USER_UPDATED',
            target_user=user,
            details={
                'username': user.username,
                'updated_by': request.user.username
            }
        )

        serializer = UserSerializer(user)
        return Response(serializer.data)

    def delete(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        AuditLog.objects.create(
            user=request.user,
            action='USER_DELETED',
            target_user=user,
            details={
                'username': user.username,
                'deleted_by': request.user.username
            }
        )

        user.delete()
        return Response({'success': True, 'message': 'User deleted successfully'})


# ============================================================
# HARDWARE MANAGEMENT VIEWS
# ============================================================

@method_decorator(csrf_exempt, name='dispatch')
class AdminHardwareCategoriesView(APIView):
    """Get all hardware categories"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        categories = HardwareCategory.objects.all()
        data = [{'id': cat.id, 'name': cat.name} for cat in categories]
        return Response(data)


@method_decorator(csrf_exempt, name='dispatch')
class AdminHardwareListView(APIView):
    """List all hardware items with filters"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        status_filter = request.query_params.get('status')
        category_filter = request.query_params.get('category')
        search = request.query_params.get('search')
        
        hardware = HardwareItem.objects.all()
        
        if status_filter:
            hardware = hardware.filter(status=status_filter)
        if category_filter:
            hardware = hardware.filter(category_id=category_filter)
        if search:
            hardware = hardware.filter(
                Q(name__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(brand__icontains=search) |
                Q(model__icontains=search)
            )
        
        hardware = hardware.order_by('-created_at')
        
        # Serialize manually with needed fields
        data = []
        for item in hardware:
            data.append({
                'id': item.id,
                'name': item.name,
                'serial_number': item.serial_number or 'N/A',
                'brand': item.brand or 'N/A',
                'model': item.model or 'N/A',
                'status': item.status,
                'condition': item.condition,
                'category_name': item.category.name if item.category else None,
                'assigned_to_name': item.assigned_to.get_full_name() if item.assigned_to else None,
                'assigned_to': item.assigned_to.id if item.assigned_to else None,
                'location': item.location,
                'purchase_date': item.purchase_date,
                'purchase_price': str(item.purchase_price) if item.purchase_price else None,
                'warranty_expiry': item.warranty_expiry,
                'supplier': item.supplier,
                'invoice_number': item.invoice_number,
                'notes': item.notes,
                'specification': item.specification,
                'created_at': item.created_at,
                'updated_at': item.updated_at,
            })
        
        return Response({
            'results': data,
            'count': hardware.count()
        })


@method_decorator(csrf_exempt, name='dispatch')
class AdminHardwareCreateView(APIView):
    """Create new hardware item"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        data = request.data
        
        # Validate required fields
        if not data.get('name'):
            return Response(
                {'error': 'Item name is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            hardware = HardwareItem.objects.create(
                name=data.get('name'),
                serial_number=data.get('serial_number', ''),
                brand=data.get('brand', ''),
                model=data.get('model', ''),
                status=data.get('status', 'available'),
                condition=data.get('condition', 'good'),
                purchase_date=data.get('purchase_date') or None,
                purchase_price=data.get('purchase_price', 0),
                warranty_expiry=data.get('warranty_expiry') or None,
                supplier=data.get('supplier', ''),
                invoice_number=data.get('invoice_number', ''),
                location=data.get('location', ''),
                notes=data.get('notes', ''),
                specification=data.get('specification', {}),
            )

            if data.get('category'):
                try:
                    category = HardwareCategory.objects.get(id=data['category'])
                    hardware.category = category
                except HardwareCategory.DoesNotExist:
                    pass

            if data.get('assigned_to'):
                try:
                    assigned_user = User.objects.get(id=data['assigned_to'])
                    hardware.assigned_to = assigned_user
                    hardware.assigned_date = timezone.now().date()
                    hardware.status = 'assigned'
                except User.DoesNotExist:
                    pass

            hardware.save()

            # Create inventory movement record
            if data.get('assigned_to'):
                InventoryMovement.objects.create(
                    hardware=hardware,
                    movement_type='assigned',
                    to_user_id=data['assigned_to'],
                    performed_by=request.user,
                    notes='Initial assignment'
                )

            AuditLog.objects.create(
                user=request.user,
                action='INVENTORY_ADDED',
                details={
                    'item_name': hardware.name,
                    'serial_number': hardware.serial_number,
                    'created_by': request.user.username
                }
            )

            return Response({
                'success': True,
                'message': 'Hardware item created successfully',
                'id': hardware.id
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


@method_decorator(csrf_exempt, name='dispatch')
class AdminHardwareDetailView(APIView):
    """Get, update, delete hardware item"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            item = HardwareItem.objects.get(pk=pk)
            data = {
                'id': item.id,
                'name': item.name,
                'serial_number': item.serial_number or 'N/A',
                'brand': item.brand or 'N/A',
                'model': item.model or 'N/A',
                'status': item.status,
                'condition': item.condition,
                'category': item.category.id if item.category else None,
                'category_name': item.category.name if item.category else None,
                'assigned_to': item.assigned_to.id if item.assigned_to else None,
                'assigned_to_name': item.assigned_to.get_full_name() if item.assigned_to else None,
                'assigned_date': item.assigned_date,
                'purchase_date': item.purchase_date,
                'purchase_price': str(item.purchase_price) if item.purchase_price else None,
                'warranty_expiry': item.warranty_expiry,
                'supplier': item.supplier,
                'invoice_number': item.invoice_number,
                'location': item.location,
                'notes': item.notes,
                'specification': item.specification,
                'created_at': item.created_at,
                'updated_at': item.updated_at,
            }
            return Response(data)
        except HardwareItem.DoesNotExist:
            return Response(
                {'error': 'Hardware item not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    def patch(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            item = HardwareItem.objects.get(pk=pk)
        except HardwareItem.DoesNotExist:
            return Response(
                {'error': 'Hardware item not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        data = request.data
        changes = []

        # Update basic fields
        for field in ['name', 'serial_number', 'brand', 'model', 'condition', 
                      'purchase_date', 'purchase_price', 'warranty_expiry', 
                      'supplier', 'invoice_number', 'location', 'notes', 'specification']:
            if field in data and data[field] != getattr(item, field):
                old_val = getattr(item, field)
                changes.append(f"{field}: {old_val} -> {data[field]}")
                setattr(item, field, data[field])

        # Update status
        if 'status' in data and data['status'] != item.status:
            changes.append(f"Status: {item.get_status_display()} -> {data['status']}")
            item.status = data['status']

        # Update category
        if 'category' in data:
            try:
                category = HardwareCategory.objects.get(id=data['category'])
                if item.category != category:
                    changes.append(f"Category: {item.category.name if item.category else 'None'} -> {category.name}")
                    item.category = category
            except HardwareCategory.DoesNotExist:
                pass

        # Update assigned_to
        if 'assigned_to' in data:
            old_assignee = item.assigned_to
            if data['assigned_to']:
                try:
                    assigned_user = User.objects.get(id=data['assigned_to'])
                    if item.assigned_to != assigned_user:
                        changes.append(f"Assigned To: {item.assigned_to.get_full_name() if item.assigned_to else 'Unassigned'} -> {assigned_user.get_full_name()}")
                        item.assigned_to = assigned_user
                        item.assigned_date = timezone.now().date()
                        if item.status == 'available':
                            item.status = 'assigned'
                except User.DoesNotExist:
                    pass
            else:
                if item.assigned_to:
                    changes.append(f"Assigned To: {item.assigned_to.get_full_name()} -> Unassigned")
                    item.assigned_to = None
                    item.assigned_date = None

        item.save()

        # Log movement if assignment changed
        if 'assigned_to' in data and changes:
            if data['assigned_to']:
                InventoryMovement.objects.create(
                    hardware=item,
                    movement_type='assigned',
                    from_user=old_assignee,
                    to_user_id=data['assigned_to'],
                    performed_by=request.user,
                    notes='Assignment updated'
                )
            else:
                InventoryMovement.objects.create(
                    hardware=item,
                    movement_type='returned',
                    from_user=old_assignee,
                    performed_by=request.user,
                    notes='Unassigned'
                )

        if changes:
            AuditLog.objects.create(
                user=request.user,
                action='INVENTORY_UPDATED',
                details={
                    'item_name': item.name,
                    'serial_number': item.serial_number,
                    'changes': ', '.join(changes),
                    'updated_by': request.user.username
                }
            )

        return Response({
            'success': True,
            'message': 'Hardware item updated successfully'
        })

    def delete(self, request, pk):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            item = HardwareItem.objects.get(pk=pk)
        except HardwareItem.DoesNotExist:
            return Response(
                {'error': 'Hardware item not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        AuditLog.objects.create(
            user=request.user,
            action='INVENTORY_DELETED',
            details={
                'item_name': item.name,
                'serial_number': item.serial_number,
                'deleted_by': request.user.username
            }
        )

        item.delete()
        return Response({
            'success': True,
            'message': 'Hardware item deleted successfully'
        })


@method_decorator(csrf_exempt, name='dispatch')
class AdminHardwareMovementsView(APIView):
    """Get hardware movements"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        movements = InventoryMovement.objects.all().order_by('-movement_date')[:20]
        data = []
        for movement in movements:
            data.append({
                'id': movement.id,
                'hardware_name': movement.hardware.name if movement.hardware else None,
                'movement_type': movement.movement_type,
                'from_user_name': movement.from_user.get_full_name() if movement.from_user else 'System',
                'to_user_name': movement.to_user.get_full_name() if movement.to_user else 'System',
                'movement_date': movement.movement_date,
                'notes': movement.notes,
            })
        
        return Response(data)


@method_decorator(csrf_exempt, name='dispatch')
class AdminHardwareReportView(APIView):
    """Generate hardware inventory report"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        hardware = HardwareItem.objects.all()
        
        total = hardware.count()
        available = hardware.filter(status='available').count()
        assigned = hardware.filter(status='assigned').count()
        maintenance = hardware.filter(status='maintenance').count()
        repair = hardware.filter(status='repair').count()
        retired = hardware.filter(status='retired').count()
        lost = hardware.filter(status='lost').count()
        
        status_breakdown = {
            'available': available,
            'assigned': assigned,
            'maintenance': maintenance,
            'repair': repair,
            'retired': retired,
            'lost': lost,
        }
        
        category_breakdown = {}
        categories = HardwareCategory.objects.all()
        for cat in categories:
            count = hardware.filter(category=cat).count()
            if count > 0:
                category_breakdown[cat.name] = count
        
        recent_movements = InventoryMovement.objects.all().order_by('-movement_date')[:10]
        movements_data = []
        for movement in recent_movements:
            movements_data.append({
                'id': movement.id,
                'hardware_name': movement.hardware.name if movement.hardware else None,
                'movement_type': movement.movement_type,
                'movement_date': movement.movement_date,
            })
        
        return Response({
            'total': total,
            'available': available,
            'assigned': assigned,
            'lost': lost,
            'status_breakdown': status_breakdown,
            'category_breakdown': category_breakdown,
            'recent_movements': movements_data,
        })


@method_decorator(csrf_exempt, name='dispatch')
class AdminHardwareExportView(APIView):
    """Export hardware inventory to Excel"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            
            hardware = HardwareItem.objects.all().order_by('name')
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hardware Inventory"
            
            # Headers
            headers = ['ID', 'Name', 'Serial Number', 'Brand', 'Model', 'Category', 
                      'Status', 'Condition', 'Assigned To', 'Location', 
                      'Purchase Date', 'Purchase Price', 'Warranty Expiry', 
                      'Supplier', 'Invoice Number', 'Created At']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row, item in enumerate(hardware, 2):
                ws.cell(row=row, column=1, value=item.id)
                ws.cell(row=row, column=2, value=item.name)
                ws.cell(row=row, column=3, value=item.serial_number or '')
                ws.cell(row=row, column=4, value=item.brand or '')
                ws.cell(row=row, column=5, value=item.model or '')
                ws.cell(row=row, column=6, value=item.category.name if item.category else '')
                ws.cell(row=row, column=7, value=item.get_status_display())
                ws.cell(row=row, column=8, value=item.get_condition_display())
                ws.cell(row=row, column=9, value=item.assigned_to.get_full_name() if item.assigned_to else '')
                ws.cell(row=row, column=10, value=item.location or '')
                ws.cell(row=row, column=11, value=item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '')
                ws.cell(row=row, column=12, value=float(item.purchase_price) if item.purchase_price else '')
                ws.cell(row=row, column=13, value=item.warranty_expiry.strftime('%Y-%m-%d') if item.warranty_expiry else '')
                ws.cell(row=row, column=14, value=item.supplier or '')
                ws.cell(row=row, column=15, value=item.invoice_number or '')
                ws.cell(row=row, column=16, value=item.created_at.strftime('%Y-%m-%d %H:%M'))
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = len(headers[col-1])
                for row in range(2, len(hardware) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="hardware_inventory_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            return Response(
                {'error': 'openpyxl library not installed. Please install it.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@method_decorator(csrf_exempt, name='dispatch')
class AdminReportSummaryView(APIView):
    """Get report summary data"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        tickets = Ticket.objects.all()
        
        total_tickets = tickets.count()
        resolved_tickets = tickets.filter(status__status_type='resolved').count()
        open_tickets = tickets.filter(status__status_type__in=['new', 'open', 'in_progress']).count()
        sla_breached = tickets.filter(sla_breached=True).count()
        
        # Ticket trend (last 7 days)
        now = timezone.now()
        ticket_trend = []
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
            count = tickets.filter(created_at__gte=day_start, created_at__lte=day_end).count()
            ticket_trend.append({
                'day': day.strftime('%a'),
                'date': day.strftime('%Y-%m-%d'),
                'count': count
            })
        
        # Status breakdown
        status_breakdown = []
        statuses = TicketStatus.objects.filter(is_active=True)
        for status in statuses:
            count = tickets.filter(status=status).count()
            if count > 0:
                status_breakdown.append({
                    'name': status.status_type,
                    'count': count
                })
        
        # Category breakdown
        category_breakdown = []
        categories = TicketCategory.objects.filter(is_active=True)
        for cat in categories:
            count = tickets.filter(category=cat).count()
            if count > 0:
                category_breakdown.append({
                    'name': cat.name,
                    'count': count
                })
        
        # SLA breakdown
        in_sla = tickets.filter(sla_breached=False, status__status_type__in=['new', 'open', 'in_progress']).count()
        at_risk = tickets.filter(sla_breached=False, status__status_type__in=['pending_customer', 'pending_third_party']).count()
        breached = tickets.filter(sla_breached=True).count()
        
        sla_breakdown = [
            {'name': 'In SLA', 'count': in_sla},
            {'name': 'At Risk', 'count': at_risk},
            {'name': 'Breached', 'count': breached},
        ]
        
        return Response({
            'total_tickets': total_tickets,
            'resolved_tickets': resolved_tickets,
            'open_tickets': open_tickets,
            'sla_breached': sla_breached,
            'ticket_trend': ticket_trend,
            'status_breakdown': status_breakdown,
            'category_breakdown': category_breakdown,
            'sla_breakdown': sla_breakdown,
        })


@method_decorator(csrf_exempt, name='dispatch')
class AdminAuditLogView(APIView):
    """Get audit logs"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        action_filter = request.query_params.get('action')
        days = int(request.query_params.get('days', 30))
        
        logs = AuditLog.objects.all()
        
        if action_filter:
            logs = logs.filter(action=action_filter)
        
        # Filter by date range
        if days:
            cutoff = timezone.now() - timedelta(days=days)
            logs = logs.filter(created_at__gte=cutoff)
        
        logs = logs.order_by('-created_at')
        
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 50))
        start = (page - 1) * page_size
        end = start + page_size
        
        paginated_logs = logs[start:end]
        
        data = []
        for log in paginated_logs:
            data.append({
                'id': log.id,
                'user': {
                    'id': log.user.id if log.user else None,
                    'username': log.user.username if log.user else 'System'
                } if log.user else None,
                'action': log.action,
                'ticket': {
                    'id': log.ticket.id if log.ticket else None,
                    'ticket_id': log.ticket.ticket_id if log.ticket else None
                } if log.ticket else None,
                'target_user': {
                    'id': log.target_user.id if log.target_user else None,
                    'username': log.target_user.username if log.target_user else None
                } if log.target_user else None,
                'details': log.details,
                'created_at': log.created_at,
            })
        
        return Response({
            'results': data,
            'count': logs.count(),
            'page': page,
            'page_size': page_size
        })
        
@method_decorator(csrf_exempt, name='dispatch')
class AdminReportGenerateView(APIView):
    """Generate a report"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        report_type = request.data.get('report_type')
        date_range = int(request.data.get('date_range', 30))

        if not report_type:
            return Response(
                {'error': 'Report type is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get the same data as summary but for the selected range
        tickets = Ticket.objects.all()
        if date_range:
            cutoff = timezone.now() - timedelta(days=date_range)
            tickets = tickets.filter(created_at__gte=cutoff)

        total_tickets = tickets.count()
        resolved_tickets = tickets.filter(status__status_type='resolved').count()
        open_tickets = tickets.filter(status__status_type__in=['new', 'open', 'in_progress']).count()
        sla_breached = tickets.filter(sla_breached=True).count()

        # Status breakdown
        status_breakdown = []
        statuses = TicketStatus.objects.filter(is_active=True)
        for status in statuses:
            count = tickets.filter(status=status).count()
            if count > 0:
                status_breakdown.append({
                    'name': status.status_type,
                    'count': count
                })

        # Category breakdown
        category_breakdown = []
        categories = TicketCategory.objects.filter(is_active=True)
        for cat in categories:
            count = tickets.filter(category=cat).count()
            if count > 0:
                category_breakdown.append({
                    'name': cat.name,
                    'count': count
                })

        # SLA breakdown
        in_sla = tickets.filter(sla_breached=False, status__status_type__in=['new', 'open', 'in_progress']).count()
        at_risk = tickets.filter(sla_breached=False, status__status_type__in=['pending_customer', 'pending_third_party']).count()
        breached = tickets.filter(sla_breached=True).count()

        sla_breakdown = [
            {'name': 'In SLA', 'count': in_sla},
            {'name': 'At Risk', 'count': at_risk},
            {'name': 'Breached', 'count': breached},
        ]

        # Ticket trend
        now = timezone.now()
        ticket_trend = []
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
            count = tickets.filter(created_at__gte=day_start, created_at__lte=day_end).count()
            ticket_trend.append({
                'day': day.strftime('%a'),
                'date': day.strftime('%Y-%m-%d'),
                'count': count
            })

        return Response({
            'total_tickets': total_tickets,
            'resolved_tickets': resolved_tickets,
            'open_tickets': open_tickets,
            'sla_breached': sla_breached,
            'status_breakdown': status_breakdown,
            'category_breakdown': category_breakdown,
            'sla_breakdown': sla_breakdown,
            'ticket_trend': ticket_trend,
        })


@method_decorator(csrf_exempt, name='dispatch')
class AdminReportExportView(APIView):
    """Export a report as Excel or PDF"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_admin(request.user):
            return Response(
                {'error': 'Access denied. Admin privileges required.'},
                status=status.HTTP_403_FORBIDDEN
            )

        report_type = request.data.get('report_type')
        date_range = int(request.data.get('date_range', 30))
        format_type = request.data.get('format', 'excel')

        if not report_type:
            return Response(
                {'error': 'Report type is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            
            tickets = Ticket.objects.all()
            if date_range:
                cutoff = timezone.now() - timedelta(days=date_range)
                tickets = tickets.filter(created_at__gte=cutoff)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"

            # Headers
            headers = ['ID', 'Title', 'Status', 'Priority', 'Category', 'Requestor', 'Created']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # Data rows
            for row, ticket in enumerate(tickets, 2):
                ws.cell(row=row, column=1, value=ticket.ticket_id)
                ws.cell(row=row, column=2, value=ticket.title)
                ws.cell(row=row, column=3, value=ticket.status.name if ticket.status else '')
                ws.cell(row=row, column=4, value=ticket.priority.name if ticket.priority else '')
                ws.cell(row=row, column=5, value=ticket.category.name if ticket.category else '')
                ws.cell(row=row, column=6, value=ticket.requestor.get_full_name() or ticket.requestor.username)
                ws.cell(row=row, column=7, value=ticket.created_at.strftime('%Y-%m-%d %H:%M'))

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = len(headers[col-1])
                for row in range(2, tickets.count() + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="report_{report_type}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            wb.save(response)
            return response

        except ImportError:
            return Response(
                {'error': 'openpyxl library not installed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )